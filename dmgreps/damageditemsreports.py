import os
from datetime import date
from openpyxl import Workbook, load_workbook
import time
import random

damage_reports = [] # Array containing all damage reports including report number, item, chair number (if applicable), location, description, informant, date, and status
excel_file = "damaged_items_reports.xlsx" # Excel file name where the reports are saved

def save_report_to_excel(report, filename=excel_file):  # Function to save a single damage report to the excel_file variable
    if os.path.exists(filename):                        # If file exists, load and append; else create new
        wb = load_workbook(filename)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Report Number", "Item", "Chair Number", "Location", "Description", "Informant", "Date", "Status"])
    
    sheet.append([  # Append the damage report data as a new row
        report["ReportNumber"],
        report["Item"],
        report.get("ChairNumber", "N/A"),
        report["Location"],
        report["Description"],
        report["Informant"],
        report["Date"],
        report["Status"]
    ])

    try:
        wb.save(filename)
        print(f"\nReport saved to {filename}\n")
    except PermissionError:
        print("\nPlease close the excel file and try again.\n") # Handle file access issues meaning this error occurs if the file is open

def report_damage(): # Function that makes a new damage report
    print("\n--- Report Damaged Items ---")
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        sheet = wb.active
        last_report_number = 0  # Find the last report number in the file (skip header if no reports exist yet)
        for row in sheet.iter_rows(min_row=2, values_only=True): # Loop through rows starting from the second row
            if row[0] is not None and isinstance(row[0], int): # Ensure the report number is valid
                last_report_number = max(last_report_number, row[0]) 
        report_number = last_report_number + 1
    else:
        report_number = 1

    item = input("Damaged Item Name: ")
    chair_number = None
    if item.lower() == "chair": # If the item is a chair, ask for chair number
        chair_number = input("Chair Number: ")
    
    location = input("Location: ")
    description = input("Damage Description: ")
    informant = input("Informant: ")
    report_date = str(date.today())

    record = { # Create the damage report record
        "ReportNumber": report_number,
        "Item": item,
        "ChairNumber": chair_number if chair_number else "N/A",
        "Location": location,
        "Description": description,
        "Informant": informant,
        "Date": report_date,
        "Status": "Pending"
    }
    print(f"\nReport Submitted Successfully! Status set to 'Pending'.")
    save_report_to_excel(record, filename=excel_file) # Save the report to the excel file using record as data and filename as the file name of the excel file

def update_status(): # Function to update the status of an existing report
    print("\n--- Update Report Status ---")
    filename = excel_file
    if not os.path.exists(filename):
        print("No reports found in damaged_items_reports.xlsx.\n")
        return

    reports = load_reports_from_excel(filename) # Load and display reports from Excel
    if len(reports) == 0:
        print("No reports found.\n")
        return

    print("\n=== Classroom Damaged Items Report ===")
    for report in reports: # Display each report with details
        print(f"\nReport {report['ReportNumber']}:")
        print(f" Item: {report['Item']}")
        if report['Item'].lower() == "chair":
            print(f" Chair Number: {report['ChairNumber']}")
        print(f" Location: {report['Location']}")
        print(f" Description: {report['Description']}")
        print(f" Reported by: {report['Informant']}")
        print(f" Date: {report['Date']}")
        print(f" Status: {report['Status']}")
    print("\nTotal Damaged Items Reported:", len(reports), "\n")

    try: 
        report_num = int(input("Enter report number to update status: ")) # Get report number to update
        if any(r["ReportNumber"] == report_num for r in reports): # Check report number exists
            new_status = input("Enter New Status (Fixed / Pending / Follow-up): ").capitalize()
            if new_status in ["Fixed", "Pending", "Follow-up"]:
                wb = load_workbook(filename)
                sheet = wb.active
                updated = False
                for row in sheet.iter_rows(min_row=2): # Find the report and update its status
                    if row[0].value == report_num: 
                        row[7].value = new_status 
                        updated = True
                        break
                if updated:
                    try:
                        wb.save(filename) 
                        print(f"\nReport {report_num} status updated to {new_status} in {filename}.\n")
                    except PermissionError: 
                        print("\nPlease close the excel file and try again.\n")
                else:
                    print("Report not found in Excel file.")
            else:
                print("Invalid Status. Please Enter 'Fixed', 'Pending', or 'Follow-up'.")
        else:
            print("Invalid Report Number.")
    except ValueError:
        print("Please Enter a Valid Number.")

def deduplicate_reports(reports): # Remove duplicates from reports based on item type, location, chair number, and date
    deduped = [] # List to hold deduplicated reports
    seen = {} # Dictionary to track reports that have been seen based on item type, location, chair number, and date
    for report in reports:
        key = (
            report.get("item_type"),
            report.get("location"),
            report.get("chair_number"),
            report.get("date"),
        )
        if key in seen: # If duplicate found, update status to the latest
            seen[key]["status"] = report.get("status", seen[key]["status"]) # Optionally, merge other fields or keep latest status
        else:
            seen[key] = report.copy() 
    deduped = list(seen.values()) # Convert the seen dictionary values back to a list
    return deduped

def generate_excel_with_duplicates(filename=excel_file, num_reports=50, duplicate_ratio=0.3): # Generate simulated data with duplicates and append to Excel if it already has reports
    if os.path.exists(filename): # If file exists, load and append; else create new
        wb = load_workbook(filename)
        sheet = wb.active
        next_report_number = sheet.max_row # max_row includes header, so subtract 1 for actual last report number
        if next_report_number > 1:
            next_report_number = sheet.max_row  # next available row for new report
        else:
            next_report_number = 1
    else: # Create new workbook and add header
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Report Number", "Item", "Chair Number", "Location", "Description", "Informant", "Date", "Status"])
        next_report_number = 1

    items = ["Chair", "Table", "Window", "Door"] # Sample items for generating items
    locations = ["Room 101", "Room 102", "Room 103"] # Sample locations for generating locations
    descriptions = ["Broken leg", "Cracked surface", "Loose hinge", "Scratched"] # Sample descriptions for generating descriptions
    informants = ["Alice", "Bob", "Charlie", "Dana"] # Sample informants for generating informants
    statuses = ["Pending", "Fixed", "Follow-up"] # Sample statuses for generating statuses

    base_reports = [] # List to hold reports that are not duplicates BEFORE adding duplicates
    for i in range(int(num_reports * (1 - duplicate_ratio))): # Generate reports without duplicates first
        item = random.choice(items)
        chair_number = str(random.randint(1, 30)) if item == "Chair" else "N/A"
        report = [
            next_report_number + i,
            item,
            chair_number,
            random.choice(locations),
            random.choice(descriptions),
            random.choice(informants),
            str(date.today()),
            random.choice(statuses)
        ]
        base_reports.append(report)
        sheet.append(report)

    for i in range(int(num_reports * duplicate_ratio)): # Add duplicates by randomly selecting from base_reports
        dup = random.choice(base_reports)
        dup_copy = dup.copy()
        dup_copy[0] = next_report_number + len(base_reports) + i  # new report number
        dup_copy[7] = random.choice(statuses)
        sheet.append(dup_copy)

    wb.save(filename) # Save the workbook with new reports
    print(f"Appended {num_reports} reports (with duplicates) to: {filename}")

def load_reports_from_excel(filename): # Function to load reports from an Excel file
    if not os.path.exists(filename):
        print(f"File {filename} not found.")
        return []
    wb = load_workbook(filename)
    sheet = wb.active
    reports = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        reports.append({
            "ReportNumber": row[0],
            "Item": row[1],
            "ChairNumber": row[2],
            "Location": row[3],
            "Description": row[4],
            "Informant": row[5],
            "Date": row[6],
            "Status": row[7]
        })
    return reports

def manual_review(filename=excel_file): # Review reports manually, measure time taken, and return summary statistics
    start = time.time() # Start timer
    reports = load_reports_from_excel(filename) # Load reports from Excel
    total_reports = len(reports) 
    items_count = {} 
    chair_numbers = set()
    for report in reports: # Count items and collect chair numbers
        item = report["Item"] 
        items_count[item] = items_count.get(item, 0) + 1 
        if item.lower() == "chair" and report["ChairNumber"] != "N/A": # If item is chair, add chair number to set
            chair_numbers.add(report["ChairNumber"])
    elapsed = time.time() - start # End timer
    print(f"\nManual review completed in {elapsed:.4f} seconds.")
    print(f"Total reports: {total_reports}")
    print("Items breakdown:", items_count)
    print("Damaged chair numbers:", ", ".join(chair_numbers) if chair_numbers else "None")
    return {
        "total_reports": total_reports,
        "items_count": items_count,
        "chair_numbers": chair_numbers,
        "elapsed": elapsed,
        "reports": reports
    }

def deduplication_review(filename=excel_file): # function to review reports using deduplication, measure time taken, and return summary statistics
    start = time.time() # Start timer
    reports = load_reports_from_excel(filename) # Load reports from Excel
    dedup_input = []
    for r in reports:
        dedup_input.append({
            "item_type": r["Item"],
            "location": r["Location"],
            "chair_number": r["ChairNumber"],
            "date": r["Date"],
            "status": r["Status"]
        })
    deduped = deduplicate_reports(dedup_input) # Deduplicate the reports
    total_reports = len(deduped)
    items_count = {}
    chair_numbers = set()
    for report in deduped: 
        item = report["item_type"]
        items_count[item] = items_count.get(item, 0) + 1
        if item.lower() == "chair" and report["chair_number"] != "N/A": # If item is chair, add chair number to set
            chair_numbers.add(report["chair_number"])
    elapsed = time.time() - start # End timer
    removed_duplicates = len(reports) - len(deduped) # Calculate number of duplicates removed
    print(f"\nDeduplication review completed in {elapsed:.4f} seconds.")
    print(f"Total reports after deduplication: {total_reports}")
    print(f"Duplicates removed: {removed_duplicates}")
    print("Items breakdown:", items_count)
    print("Damaged chair numbers:", ", ".join(chair_numbers) if chair_numbers else "None")
    return {
        "total_reports": total_reports,
        "items_count": items_count,
        "chair_numbers": chair_numbers,
        "elapsed": elapsed,
        "removed_duplicates": removed_duplicates,
        "reports": deduped
    }

def compare_summary_statistics(manual_stats, dedup_stats): # Function to compare summary statistics before and after deduplication
    """
    Compare summary statistics before and after deduplication.
    """
    print("\n=== Accuracy of Summary Statistics ===")
    print(f"Manual total reports: {manual_stats['total_reports']}")
    print(f"Deduplicated total reports: {dedup_stats['total_reports']}")
    print(f"Duplicates removed: {dedup_stats['removed_duplicates']}")
    print("\nManual items breakdown:", manual_stats['items_count'])
    print("Deduplicated items breakdown:", dedup_stats['items_count'])
    print("\nManual damaged chair numbers:", ", ".join(manual_stats['chair_numbers']) if manual_stats['chair_numbers'] else "None")
    print("Deduplicated damaged chair numbers:", ", ".join(dedup_stats['chair_numbers']) if dedup_stats['chair_numbers'] else "None")
    print("\nManual review time: {:.4f} seconds".format(manual_stats['elapsed']))
    print("Deduplication review time: {:.4f} seconds".format(dedup_stats['elapsed']))
    print("\n======================\n")

def main():
    while True:
        print("=== Damaged Items Report System ===")
        print("1. Report Damaged Item")
        print("2. Update Report Status")
        print("3. Generate Simulated Data")
        print("4. Generate Summary and Compare Statistics")
        print("5. Exit")
        choice = input(f"\nEnter Your Choice: ")

        if choice == "1":
            report_damage()
        elif choice == "2":
            update_status()
        elif choice == "3":
            generate_excel_with_duplicates()
        elif choice == "4":
            manual_stats = manual_review()
            dedup_stats = deduplication_review()
            compare_summary_statistics(manual_stats, dedup_stats)
        elif choice == "5":
            break
        else:
            print("Invalid choice, please try again.\n")

if __name__ == "__main__": main()
